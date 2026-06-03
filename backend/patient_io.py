"""Excel/CSV import/export for patients (spreadsheet column layout)."""
from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from saas import iso, now_utc

EXPORT_COLUMNS = [
    "FirstName",
    "LastName",
    "Phone No",
    "UserCode",
    "membershipname",
    "lastvisit",
    "guestIconInformation",
]


def _norm_header(h: str) -> str:
    s = (h or "").replace("\ufeff", "").strip()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


HEADER_ALIASES = {
    "firstname": "first_name",
    "first name": "first_name",
    "lastname": "last_name",
    "last name": "last_name",
    "full name": "full_name",
    "fullname": "full_name",
    "name": "full_name",
    "phone no": "phone",
    "phone": "phone",
    "phoneno": "phone",
    "phone number": "phone",
    "mobile": "phone",
    "usercode": "user_code",
    "user code": "user_code",
    "membershipname": "membership_name",
    "membership name": "membership_name",
    "lastvisit": "last_visit",
    "last visit": "last_visit",
    "guesticoninformation": "guest_icon_information",
    "guest icon information": "guest_icon_information",
}


def _resolve_header(norm: str) -> Optional[str]:
    if not norm:
        return None
    if norm in HEADER_ALIASES:
        return HEADER_ALIASES[norm]
    if "first" in norm and "name" in norm:
        return "first_name"
    if norm in ("fname", "given name", "givenname", "forename"):
        return "first_name"
    if "last" in norm and "name" in norm:
        return "last_name"
    if norm in ("lname", "surname", "family name", "familyname"):
        return "last_name"
    if norm in ("patient name", "patientname", "client name"):
        return "full_name"
    if "phone" in norm or norm in ("tel", "telephone", "contact", "whatsapp"):
        return "phone"
    if "user" in norm and "code" in norm:
        return "user_code"
    if norm in ("code", "patient code", "client code", "member code"):
        return "user_code"
    if "membership" in norm:
        return "membership_name"
    if "last" in norm and "visit" in norm:
        return "last_visit"
    if "guest" in norm and "icon" in norm:
        return "guest_icon_information"
    return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _find_header_row(rows: List[tuple]) -> Tuple[int, List[Any]]:
    """Locate header row — some exports include title rows above the columns."""
    for idx, row in enumerate(rows[:25]):
        mapped = _map_columns_indexed(list(row))
        vals = set(mapped.values())
        if "first_name" in vals or "last_name" in vals or "full_name" in vals:
            return idx, list(row)
    return 0, list(rows[0]) if rows else []


def _map_columns_indexed(headers: List[Any]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for j, raw in enumerate(headers):
        internal = _resolve_header(_norm_header(_cell_str(raw)))
        if internal:
            out[j] = internal
    return out


def _headers_preview(headers: List[Any]) -> str:
    labels = [_cell_str(h) for h in headers if _cell_str(h)]
    return ", ".join(labels[:12]) if labels else "(empty header row)"


def parse_last_visit(val: Any) -> str:
    """Parse to ISO date YYYY-MM-DD. Accepts DD/MM/YYYY or ISO."""
    if val is None or str(val).strip() == "":
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return ""
    m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m2:
        return s
    return ""


def format_last_visit_export(iso_date: str) -> str:
    if not iso_date:
        return ""
    try:
        d = date.fromisoformat(iso_date[:10])
        return d.strftime("%d/%m/%Y")
    except ValueError:
        return iso_date


def split_full_name(full_name: str) -> Tuple[str, str]:
    parts = (full_name or "").strip().split(None, 1)
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def patient_to_export_row(p: dict) -> Dict[str, Any]:
    first = (p.get("first_name") or "").strip()
    last = (p.get("last_name") or "").strip()
    if not first and not last:
        first, last = split_full_name(p.get("full_name") or "")
    return {
        "FirstName": first,
        "LastName": last,
        "Phone No": p.get("phone") or "",
        "UserCode": p.get("user_code") or "",
        "membershipname": p.get("membership_name") or "",
        "lastvisit": format_last_visit_export(p.get("last_visit") or ""),
        "guestIconInformation": p.get("guest_icon_information") or "",
    }


def map_headers(fieldnames: List[str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for raw in fieldnames or []:
        internal = _resolve_header(_norm_header(str(raw) if raw is not None else ""))
        if internal:
            out[str(raw)] = internal
    return out


def _rows_from_table(all_rows: List[tuple]) -> Tuple[List[dict], List[dict]]:
    if not all_rows:
        return [], [{"row": 0, "message": "File has no rows"}]

    header_idx, headers = _find_header_row(all_rows)
    col_map = _map_columns_indexed(headers)
    if not any(k in col_map.values() for k in ("first_name", "last_name", "full_name")):
        preview = _headers_preview(headers)
        return [], [{
            "row": header_idx + 1,
            "message": f"Missing name columns (FirstName / LastName). Found: {preview}",
        }]

    data_rows = all_rows[header_idx + 1:]
    parsed: List[dict] = []
    errors: List[dict] = []
    for i, raw_row in enumerate(data_rows, start=header_idx + 2):
        if not raw_row or not any(_cell_str(c) for c in raw_row):
            continue
        data: Dict[str, Any] = {}
        row_list = list(raw_row)
        for j, internal in col_map.items():
            val = row_list[j] if j < len(row_list) else ""
            data[internal] = _cell_str(val)

        first = data.get("first_name", "").strip()
        last = data.get("last_name", "").strip()
        full = data.get("full_name", "").strip()
        if full and not first and not last:
            first, last = split_full_name(full)
        if not first and not last:
            errors.append({"row": i, "message": "FirstName or LastName is required"})
            continue

        full_name = f"{first} {last}".strip()
        parsed.append({
            "first_name": first,
            "last_name": last,
            "full_name": full_name,
            "phone": data.get("phone", "").strip(),
            "user_code": data.get("user_code", "").strip(),
            "membership_name": data.get("membership_name", "").strip(),
            "last_visit": parse_last_visit(data.get("last_visit", "")),
            "guest_icon_information": data.get("guest_icon_information", "").strip(),
        })
    return parsed, errors


def parse_csv_text(text: str) -> Tuple[List[dict], List[dict]]:
    text = text.lstrip("\ufeff")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [{"row": 0, "message": "File has no header row"}]
    return _rows_from_table([tuple(r) for r in rows])


def parse_xlsx_bytes(raw: bytes) -> Tuple[List[dict], List[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel import") from ex
    if raw[:2] != b"PK":
        return [], [{
            "row": 0,
            "message": "Not a valid .xlsx file. In Excel use Save As → Excel Workbook (.xlsx), not .xls",
        }]
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    last_errors: List[dict] = [{"row": 0, "message": "Excel workbook has no patient rows"}]
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parsed, errors = _rows_from_table(rows)
        if parsed:
            wb.close()
            return parsed, errors
        if errors:
            last_errors = errors
    wb.close()
    return [], last_errors


def rows_to_csv(rows: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def rows_to_xlsx(rows: List[dict]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel export") from ex
    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in EXPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_patient_doc(row: dict, clinic_id: str, user_id: str, existing: Optional[dict] = None) -> dict:
    doc = {
        "first_name": row.get("first_name") or "",
        "last_name": row.get("last_name") or "",
        "full_name": row["full_name"],
        "phone": row.get("phone") or "",
        "user_code": row.get("user_code") or "",
        "membership_name": row.get("membership_name") or "",
        "last_visit": row.get("last_visit") or "",
        "guest_icon_information": row.get("guest_icon_information") or "",
        "clinic_id": clinic_id,
    }
    if existing:
        doc["gender"] = existing.get("gender")
        doc["date_of_birth"] = existing.get("date_of_birth")
        doc["email"] = existing.get("email")
        doc["address"] = existing.get("address")
        doc["medical_history"] = existing.get("medical_history")
        doc["allergies"] = existing.get("allergies")
        doc["notes"] = existing.get("notes")
        doc["created_at"] = existing.get("created_at")
        doc["created_by"] = existing.get("created_by")
        doc["id"] = existing["id"]
    else:
        doc["gender"] = None
        doc["date_of_birth"] = None
        doc["email"] = None
        doc["address"] = None
        doc["medical_history"] = None
        doc["allergies"] = None
        doc["notes"] = None
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = iso(now_utc())
        doc["created_by"] = user_id
    return doc
