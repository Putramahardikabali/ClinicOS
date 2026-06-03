"""Excel export for commission records."""
from __future__ import annotations

import calendar
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# Legacy columns (kept for backward-compatible helpers/tests)
EXPORT_COLUMNS = [
    "ApprovedDate",
    "InvoiceNumber",
    "PatientName",
    "ItemName",
    "ItemType",
    "PerformerName",
    "PerformerRole",
    "GrossAmount",
    "DiscountAmount",
    "NetAmount",
    "PaidAmount",
    "CommissionRule",
    "CommissionType",
    "CommissionValue",
    "CalculationBasis",
    "CommissionAmount",
    "Status",
    "ApprovedAt",
    "PaidOutAt",
]

STAFF_SUMMARY_COLUMNS = [
    "Staff Name",
    "Role",
    "Period Start",
    "Period End",
    "Total Gross Sales",
    "Total Discount",
    "Total Net Sales",
    "Total Paid Amount",
    "Total Commission",
    "Approved Count",
    "Paid Out Count",
    "Remaining Approved Unpaid",
]

DETAIL_COLUMNS = [
    "Date",
    "Invoice Number",
    "Patient",
    "Item",
    "Gross Amount",
    "Discount",
    "Net Amount",
    "Paid Amount",
    "Rule",
    "Commission Amount",
    "Status",
    "Approved At",
    "Paid Out At",
]

DATE_BASIS_FIELDS = {
    "earned_at": "created_at",
    "approved_at": "approved_at",
    "paid_out_at": "paid_out_at",
    "invoice_paid_at": "created_at",
}


def month_to_range(month: str) -> Tuple[str, str]:
    """YYYY-MM → (first day, last day) as YYYY-MM-DD."""
    parts = (month or "").strip().split("-")
    if len(parts) != 2:
        raise ValueError("month must be YYYY-MM")
    year, mon = int(parts[0]), int(parts[1])
    if mon < 1 or mon > 12:
        raise ValueError("invalid month")
    last = calendar.monthrange(year, mon)[1]
    return f"{year}-{mon:02d}-01", f"{year}-{mon:02d}-{last:02d}"


def format_export_date(iso: str) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        return iso[:10]


def record_date_for_basis(rec: dict, date_basis: str = "approved_at") -> str:
    field = DATE_BASIS_FIELDS.get(date_basis, "approved_at")
    if field == "approved_at":
        return rec.get("approved_at") or rec.get("created_at") or ""
    if field == "paid_out_at":
        return rec.get("paid_out_at") or ""
    return rec.get("created_at") or ""


def commission_record_to_export_row(rec: dict, patient_name: str = "") -> Dict[str, Any]:
    ctype = rec.get("commission_type") or ""
    cval = rec.get("commission_value") or 0
    if ctype == "percentage":
        rate = f"{cval}%"
    elif ctype == "fixed_amount":
        rate = int(cval)
    else:
        rate = ""
    return {
        "ApprovedDate": format_export_date(rec.get("approved_at") or ""),
        "InvoiceNumber": rec.get("invoice_number_snapshot") or "",
        "PatientName": patient_name or "",
        "ItemName": rec.get("item_name_snapshot") or "",
        "ItemType": rec.get("item_type") or "",
        "PerformerName": rec.get("staff_name_snapshot") or "",
        "PerformerRole": rec.get("staff_role_snapshot") or "",
        "GrossAmount": int(rec.get("gross_amount") or 0),
        "DiscountAmount": int(rec.get("discount_amount") or 0),
        "NetAmount": int(rec.get("net_amount") or 0),
        "PaidAmount": int(rec.get("paid_amount") or 0),
        "CommissionRule": rec.get("commission_rule_name_snapshot") or "",
        "CommissionType": ctype,
        "CommissionValue": rate,
        "CalculationBasis": rec.get("calculation_basis") or "",
        "CommissionAmount": int(rec.get("commission_amount") or 0),
        "Status": rec.get("status") or "",
        "ApprovedAt": format_export_date(rec.get("approved_at") or ""),
        "PaidOutAt": format_export_date(rec.get("paid_out_at") or ""),
    }


def commission_record_to_detail_row(
    rec: dict,
    patient_name: str = "",
    *,
    date_basis: str = "approved_at",
) -> Dict[str, Any]:
    display_date = record_date_for_basis(rec, date_basis)
    return {
        "Date": format_export_date(display_date),
        "Invoice Number": rec.get("invoice_number_snapshot") or "",
        "Patient": patient_name or "",
        "Item": rec.get("item_name_snapshot") or "",
        "Gross Amount": int(rec.get("gross_amount") or 0),
        "Discount": int(rec.get("discount_amount") or 0),
        "Net Amount": int(rec.get("net_amount") or 0),
        "Paid Amount": int(rec.get("paid_amount") or 0),
        "Rule": rec.get("commission_rule_name_snapshot") or "",
        "Commission Amount": int(rec.get("commission_amount") or 0),
        "Status": rec.get("status") or "",
        "Approved At": format_export_date(rec.get("approved_at") or ""),
        "Paid Out At": format_export_date(rec.get("paid_out_at") or ""),
    }


def build_staff_period_summary(
    enriched_rows: List[dict],
    *,
    staff_name: str,
    staff_role: str,
    period_start: str,
    period_end: str,
) -> Dict[str, Any]:
    gross = discount = net = paid = commission = 0
    approved_count = paid_out_count = 0
    remaining_approved_unpaid = 0
    for rec in enriched_rows:
        gross += int(rec.get("gross_amount") or 0)
        discount += int(rec.get("discount_amount") or 0)
        net += int(rec.get("net_amount") or 0)
        paid += int(rec.get("paid_amount") or 0)
        amt = int(rec.get("commission_amount") or 0)
        commission += amt
        status = rec.get("status") or ""
        if status == "approved":
            approved_count += 1
            remaining_approved_unpaid += amt
        elif status == "paid_out":
            paid_out_count += 1
    return {
        "Staff Name": staff_name,
        "Role": staff_role,
        "Period Start": period_start,
        "Period End": period_end,
        "Total Gross Sales": gross,
        "Total Discount": discount,
        "Total Net Sales": net,
        "Total Paid Amount": paid,
        "Total Commission": commission,
        "Approved Count": approved_count,
        "Paid Out Count": paid_out_count,
        "Remaining Approved Unpaid": remaining_approved_unpaid,
    }


def rows_to_xlsx(rows: List[dict], sheet_title: str = "Commission") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel export") from ex

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(c, "") for c in EXPORT_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_staff_export_xlsx(
    summary: Dict[str, Any],
    detail_rows: List[dict],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel export") from ex

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"[:31]
    ws_summary.append(STAFF_SUMMARY_COLUMNS)
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    ws_summary.append([summary.get(c, "") for c in STAFF_SUMMARY_COLUMNS])

    ws_detail = wb.create_sheet(title="Detailed Commission Items"[:31])
    ws_detail.append(DETAIL_COLUMNS)
    for cell in ws_detail[1]:
        cell.font = Font(bold=True)
    for row in detail_rows:
        ws_detail.append([row.get(c, "") for c in DETAIL_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
