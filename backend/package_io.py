"""Excel/CSV import/export for clinic packages catalog."""
from __future__ import annotations

import csv
import io
import re
import uuid
from typing import Any, Dict, List, Optional, Tuple

from saas import iso, now_utc

EXPORT_COLUMNS = [
    "PackageName",
    "PackageCode",
    "Status",
    "PackageType",
    "PackagePrice",
    "BusinessUnit",
    "OnlineBooking",
]

PACKAGE_TYPES = ["Series package", "Day package"]


def _norm_header(h: str) -> str:
    s = (h or "").replace("\ufeff", "").strip()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


HEADER_ALIASES = {
    "packagename": "name",
    "package name": "name",
    "packagecode": "package_code",
    "package code": "package_code",
    "status": "status",
    "packagetype": "package_type",
    "package type": "package_type",
    "packageprice": "price",
    "package price": "price",
    "price": "price",
    "packagecategory": "category",
    "package category": "category",
    "businessunit": "business_unit",
    "business unit": "business_unit",
    "onlinebooking": "online_booking",
    "online booking": "online_booking",
}


def _resolve_header(norm: str) -> Optional[str]:
    if not norm:
        return None
    if norm in HEADER_ALIASES:
        return HEADER_ALIASES[norm]
    if "package" in norm and "name" in norm:
        return "name"
    if "package" in norm and "code" in norm:
        return "package_code"
    if "package" in norm and "type" in norm:
        return "package_type"
    if "package" in norm and "price" in norm:
        return "price"
    if "package" in norm and "categ" in norm:
        return "category"
    if "business" in norm and "unit" in norm:
        return "business_unit"
    if "online" in norm and "book" in norm:
        return "online_booking"
    if norm == "status":
        return "status"
    return None


def parse_price_idr(val: Any) -> int:
    if val is None or str(val).strip() == "":
        return 0
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return max(0, int(val))
    s = re.sub(r"(?i)rp\.?\s*", "", str(val).strip())
    s = s.replace(",", "").replace(" ", "")
    if s.count(".") >= 1 and len(s.split(".")[-1]) == 3:
        s = s.replace(".", "")
    try:
        return max(0, int(float(s)))
    except (TypeError, ValueError):
        return 0


def parse_bool(val: Any, default: bool = False) -> bool:
    if val is None or str(val).strip() == "":
        return default
    s = str(val).strip().lower()
    if s in ("true", "yes", "y", "1", "on", "active"):
        return True
    if s in ("false", "no", "n", "0", "off", "inactive"):
        return False
    return default


def parse_status_active(val: Any) -> bool:
    if val is None or str(val).strip() == "":
        return True
    s = str(val).strip().lower()
    if s in ("inactive", "disabled", "no"):
        return False
    return True


def format_price_export(n: int) -> str:
    n = int(n or 0)
    s = str(n)
    parts = []
    while s:
        parts.append(s[-3:])
        s = s[:-3]
    return ".".join(reversed(parts))


def package_to_export_row(p: dict) -> Dict[str, Any]:
    return {
        "PackageName": p.get("name") or "",
        "PackageCode": p.get("package_code") or p.get("key") or "",
        "Status": "Active" if p.get("active", True) else "Inactive",
        "PackageType": p.get("package_type") or "Series package",
        "PackagePrice": format_price_export(p.get("price_idr") or 0),
        "BusinessUnit": p.get("business_unit") or "Default",
        "OnlineBooking": "Yes" if p.get("online_booking") else "No",
    }


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _map_columns_indexed(headers: List[Any]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for j, raw in enumerate(headers):
        internal = _resolve_header(_norm_header(_cell_str(raw)))
        if internal:
            out[j] = internal
    return out


def _find_header_row(rows: List[tuple]) -> Tuple[int, List[Any]]:
    for idx, row in enumerate(rows[:25]):
        mapped = _map_columns_indexed(list(row))
        if "name" in mapped.values():
            return idx, list(row)
    return 0, list(rows[0]) if rows else []


def _rows_from_table(all_rows: List[tuple]) -> Tuple[List[dict], List[dict]]:
    if not all_rows:
        return [], [{"row": 0, "message": "File has no rows"}]
    header_idx, headers = _find_header_row(all_rows)
    col_map = _map_columns_indexed(headers)
    if "name" not in col_map.values():
        preview = ", ".join(_cell_str(h) for h in headers if _cell_str(h))[:120]
        return [], [{"row": header_idx + 1, "message": f"Missing PackageName column. Found: {preview or '(empty)'}"}]

    internal_fields = set(col_map.values())
    parsed: List[dict] = []
    errors: List[dict] = []
    for i, raw_row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        if not raw_row or not any(_cell_str(c) for c in raw_row):
            continue
        data: Dict[str, Any] = {}
        row_list = list(raw_row)
        for j, internal in col_map.items():
            data[internal] = row_list[j] if j < len(row_list) else ""
        name = _cell_str(data.get("name"))
        if not name:
            errors.append({"row": i, "message": "PackageName is required"})
            continue
        code = _cell_str(data.get("package_code"))
        if not code:
            code = re.sub(r"[^a-zA-Z0-9]+", "", name)[:32] or str(uuid.uuid4())[:8].upper()
        ptype = _cell_str(data.get("package_type")) or "Series package"
        row_out = {
            "name": name,
            "package_code": code,
            "key": code[:32],
            "active": parse_status_active(data.get("status")),
            "package_type": ptype,
            "price_idr": parse_price_idr(data.get("price")),
            "online_booking": parse_bool(data.get("online_booking"), False),
        }
        if "category" in internal_fields:
            row_out["category"] = _cell_str(data.get("category")) or "Default"
        if "business_unit" in internal_fields:
            row_out["business_unit"] = _cell_str(data.get("business_unit")) or "Default"
        parsed.append(row_out)
    return parsed, errors


def parse_csv_text(text: str) -> Tuple[List[dict], List[dict]]:
    text = text.lstrip("\ufeff")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [{"row": 0, "message": "File has no header row"}]
    return _rows_from_table([tuple(r) for r in rows])


def parse_xlsx_bytes(raw: bytes) -> Tuple[List[dict], List[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel import") from ex
    if raw[:2] != b"PK":
        return [], [{"row": 0, "message": "Not a valid .xlsx file. Save as Excel Workbook (.xlsx)"}]
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    last_errors: List[dict] = [{"row": 0, "message": "No package rows found"}]
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parsed, errors = _rows_from_table(rows)
        if parsed:
            wb.close()
            return parsed, errors
        if errors:
            last_errors = errors
    wb.close()
    return [], last_errors


def rows_to_csv(rows: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def rows_to_xlsx(rows: List[dict]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel export") from ex
    wb = Workbook()
    ws = wb.active
    ws.title = "Packages"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in EXPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_package_doc(row: dict, clinic_id: str, user_id: str, existing: Optional[dict] = None) -> dict:
    doc = {
        "name": row["name"],
        "package_code": row.get("package_code") or row["name"][:32],
        "key": row.get("key") or row.get("package_code", "")[:32],
        "package_type": row.get("package_type") or "Series package",
        "active": bool(row.get("active", True)),
        "online_booking": bool(row.get("online_booking", False)),
        "price_idr": int(row.get("price_idr") or 0),
        "sessions_total": int(row.get("sessions_total") or 6),
        "valid_days": int(row.get("valid_days") or 365),
        "duration_min": int(row.get("duration_min") or 60),
        "performer_type": row.get("performer_type") or "therapist",
        "description": row.get("description") or "",
        "clinic_id": clinic_id,
    }
    if "category" in row:
        doc["category"] = row.get("category") or "Default"
    elif existing:
        doc["category"] = existing.get("category") or "Default"
    else:
        doc["category"] = "Default"
    if "business_unit" in row:
        doc["business_unit"] = row.get("business_unit") or "Default"
    elif existing:
        doc["business_unit"] = existing.get("business_unit") or "Default"
    else:
        doc["business_unit"] = "Default"
    if existing:
        doc["duration_min"] = existing.get("duration_min", doc["duration_min"])
        doc["performer_type"] = existing.get("performer_type", doc["performer_type"])
        doc["description"] = existing.get("description") or ""
        doc["created_at"] = existing.get("created_at")
        doc["created_by"] = existing.get("created_by")
        doc["id"] = existing["id"]
    else:
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = iso(now_utc())
        doc["created_by"] = user_id
    return doc
