"""Excel/CSV import/export for clinic products (internal inventory)."""
from __future__ import annotations

import csv
import io
import re
import uuid
from typing import Any, Dict, List, Optional, Tuple

from package_io import format_price_export, parse_price_idr
from saas import iso, now_utc

EXPORT_COLUMNS = [
    "Product Code",
    "Product Name",
    "Brand",
    "Product Type",
    "Category",
    "Current Stock",
    "Unit",
    "Minimum Stock",
    "Active",
    "Notes",
]


def _norm_header(h: str) -> str:
    s = (h or "").replace("\ufeff", "").strip()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


HEADER_ALIASES = {
    "productcode": "product_code",
    "product code": "product_code",
    "code": "product_code",
    "productname": "name",
    "product name": "name",
    "name": "name",
    "brand": "brand",
    "category": "category",
    "sub category": "sub_category",
    "subcategory": "sub_category",
    "businessunit": "business_unit",
    "business unit": "business_unit",
    "producttype": "product_type",
    "product type": "product_type",
    "type": "product_type",
    "currentstock": "current_stock",
    "current stock": "current_stock",
    "stock": "current_stock",
    "unit": "unit",
    "minimumstock": "minimum_stock",
    "minimum stock": "minimum_stock",
    "min stock": "minimum_stock",
    "notes": "notes",
    "active": "status",
    "status": "status",
    "saleprice": "sale_price",
    "sale price": "sale_price",
    "sale price idr": "sale_price",
    "mrp": "mrp",
    "mrp idr": "mrp",
    "amount": "amount",
}


def _resolve_header(norm: str) -> Optional[str]:
    if not norm:
        return None
    if norm in HEADER_ALIASES:
        return HEADER_ALIASES[norm]
    if "product" in norm and "code" in norm:
        return "product_code"
    if "product" in norm and "name" in norm:
        return "name"
    if "product" in norm and "type" in norm:
        return "product_type"
    if "sub" in norm and "categ" in norm:
        return "sub_category"
    if "business" in norm and "unit" in norm:
        return "business_unit"
    if "minimum" in norm and "stock" in norm:
        return "minimum_stock"
    if "current" in norm and "stock" in norm:
        return "current_stock"
    if norm == "stock":
        return "current_stock"
    if norm == "notes" or norm == "note":
        return "notes"
    if norm == "unit":
        return "unit"
    if norm == "active":
        return "status"
    if "sale" in norm and "price" in norm:
        return "sale_price"
    if norm == "mrp" or ("mrp" in norm and "idr" in norm):
        return "mrp"
    return None


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def normalize_code(code: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "", (code or "")).lower()


def code_from_name(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "", name or "")
    return s[:32] if s else str(uuid.uuid4())[:8].upper()


def build_product_lookup(existing_rows: List[dict]) -> Dict[str, dict]:
    lookup: Dict[str, dict] = {}
    for t in existing_rows:
        for code in (t.get("product_code"), t.get("key")):
            nk = normalize_code(code or "")
            if nk:
                lookup[f"code:{nk}"] = t
        nm = normalize_name(t.get("name") or "")
        if nm:
            lookup[f"name:{nm}"] = t
        slug = normalize_code(t.get("name") or "")
        if slug:
            lookup[f"slug:{slug}"] = t
    return lookup


def find_product_match(row: dict, lookup: Dict[str, dict]) -> Optional[dict]:
    code = normalize_code(row.get("product_code") or "")
    if code:
        hit = lookup.get(f"code:{code}")
        if hit:
            return hit
    name = normalize_name(row.get("name") or "")
    if name:
        hit = lookup.get(f"name:{name}")
        if hit:
            return hit
    slug = normalize_code(row.get("name") or "")
    if slug:
        return lookup.get(f"slug:{slug}")
    return None


def register_product_in_lookup(lookup: Dict[str, dict], doc: dict) -> None:
    for code in (doc.get("product_code"), doc.get("key")):
        nk = normalize_code(code or "")
        if nk:
            lookup[f"code:{nk}"] = doc
    nm = normalize_name(doc.get("name") or "")
    if nm:
        lookup[f"name:{nm}"] = doc
    slug = normalize_code(doc.get("name") or "")
    if slug:
        lookup[f"slug:{slug}"] = doc


def parse_status_active(val: Any) -> bool:
    if val is None or str(val).strip() == "":
        return True
    s = str(val).strip().lower()
    if s in ("inactive", "disabled", "no"):
        return False
    return True


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def parse_int_qty(val: Any, default: int = 0) -> int:
    if val is None or str(val).strip() == "":
        return default
    try:
        return max(0, int(float(str(val).replace(",", "").strip())))
    except (TypeError, ValueError):
        return default


def product_to_export_row(p: dict) -> Dict[str, Any]:
    return {
        "Product Code": p.get("product_code") or p.get("key") or "",
        "Product Name": p.get("name") or "",
        "Brand": p.get("brand") or "",
        "Product Type": p.get("product_type") or "",
        "Category": p.get("category") or "",
        "Current Stock": int(p.get("current_stock") or 0),
        "Unit": p.get("unit") or "pcs",
        "Minimum Stock": int(p.get("minimum_stock") or 0),
        "Active": "Active" if p.get("active", True) else "Inactive",
        "Notes": p.get("notes") or "",
    }


def _map_columns_indexed(headers: List[Any]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for j, raw in enumerate(headers):
        internal = _resolve_header(_norm_header(_cell_str(raw)))
        if internal:
            out[j] = internal
    return out


def _find_header_row(rows: List[tuple]) -> Tuple[int, List[Any]]:
    for idx, row in enumerate(rows[:25]):
        mapped = _map_columns_indexed(list(row))
        if "name" in mapped.values():
            return idx, list(row)
    return 0, list(rows[0]) if rows else []


def _rows_from_table(all_rows: List[tuple]) -> Tuple[List[dict], List[dict]]:
    if not all_rows:
        return [], [{"row": 0, "message": "File has no rows"}]
    header_idx, headers = _find_header_row(all_rows)
    col_map = _map_columns_indexed(headers)
    if "name" not in col_map.values():
        preview = ", ".join(_cell_str(h) for h in headers if _cell_str(h))[:120]
        return [], [{"row": header_idx + 1, "message": f"Missing Product Name column. Found: {preview or '(empty)'}"}]
    has_sale = "sale_price" in col_map.values()
    has_mrp = "mrp" in col_map.values()
    has_current_stock = "current_stock" in col_map.values()
    has_minimum_stock = "minimum_stock" in col_map.values()

    parsed: List[dict] = []
    errors: List[dict] = []
    for i, raw_row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        if not raw_row or not any(_cell_str(c) for c in raw_row):
            continue
        data: Dict[str, Any] = {}
        row_list = list(raw_row)
        for j, internal in col_map.items():
            data[internal] = row_list[j] if j < len(row_list) else ""
        name = _cell_str(data.get("name"))
        if not name:
            errors.append({"row": i, "message": "Product Name is required"})
            continue
        raw_code = _cell_str(data.get("product_code"))
        product_code = raw_code or code_from_name(name)
        row_out = {
            "name": name,
            "product_code": product_code,
            "product_code_provided": bool(raw_code),
            "brand": _cell_str(data.get("brand")),
            "category": _cell_str(data.get("category")) or "Default",
            "sub_category": _cell_str(data.get("sub_category")),
            "business_unit": _cell_str(data.get("business_unit")) or "Default",
            "product_type": _cell_str(data.get("product_type")) or "Consumable",
            "amount": _cell_str(data.get("amount")),
            "unit": _cell_str(data.get("unit")) or "pcs",
            "notes": _cell_str(data.get("notes")),
            "active": parse_status_active(data.get("status")),
        }
        if has_current_stock:
            row_out["current_stock"] = parse_int_qty(data.get("current_stock"))
        if has_minimum_stock:
            row_out["minimum_stock"] = parse_int_qty(data.get("minimum_stock"))
        if has_sale:
            row_out["sale_price_idr"] = parse_price_idr(data.get("sale_price"))
        if has_mrp:
            row_out["mrp_idr"] = parse_price_idr(data.get("mrp"))
        parsed.append(row_out)
    return parsed, errors


def parse_csv_text(text: str) -> Tuple[List[dict], List[dict]]:
    text = text.lstrip("\ufeff")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [{"row": 0, "message": "File has no header row"}]
    return _rows_from_table([tuple(r) for r in rows])


def parse_xlsx_bytes(raw: bytes) -> Tuple[List[dict], List[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel import") from ex
    if raw[:2] != b"PK":
        return [], [{"row": 0, "message": "Not a valid .xlsx file. Save as Excel Workbook (.xlsx)"}]
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    last_errors: List[dict] = [{"row": 0, "message": "No product rows found"}]
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parsed, errors = _rows_from_table(rows)
        if parsed:
            wb.close()
            return parsed, errors
        if errors:
            last_errors = errors
    wb.close()
    return [], last_errors


def rows_to_csv(rows: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def rows_to_xlsx(rows: List[dict]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel export") from ex
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in EXPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_product_doc(row: dict, clinic_id: str, user_id: str, existing: Optional[dict] = None) -> dict:
    if existing and not row.get("product_code_provided"):
        product_code = (existing.get("product_code") or row.get("product_code") or "").strip()
    else:
        product_code = (row.get("product_code") or "").strip()
    if not product_code:
        product_code = code_from_name(row["name"])
    key = (existing.get("key") if existing else None) or product_code[:32]
    doc = {
        "name": row["name"],
        "product_code": product_code,
        "key": key,
        "brand": row.get("brand") or "",
        "category": row.get("category") or "Default",
        "sub_category": row.get("sub_category") or "",
        "business_unit": row.get("business_unit") or "Default",
        "product_type": row.get("product_type") or "Consumable",
        "amount": row.get("amount") or "",
        "active": bool(row.get("active", True)),
        "clinic_id": clinic_id,
        "notes": row.get("notes") or (existing.get("notes") if existing else "") or "",
        "unit": row.get("unit") or (existing.get("unit") if existing else "") or "pcs",
    }
    if "current_stock" in row:
        doc["current_stock"] = int(row.get("current_stock") or 0)
    elif existing:
        doc["current_stock"] = int(existing.get("current_stock") or 0)
    else:
        doc["current_stock"] = 0
    if "minimum_stock" in row:
        doc["minimum_stock"] = int(row.get("minimum_stock") or 0)
    elif existing:
        doc["minimum_stock"] = int(existing.get("minimum_stock") or 0)
    else:
        doc["minimum_stock"] = 0
    doc["stock_updated_at"] = iso(now_utc())
    if "sale_price_idr" in row:
        doc["sale_price_idr"] = int(row.get("sale_price_idr") or 0)
    elif existing:
        doc["sale_price_idr"] = int(existing.get("sale_price_idr") or 0)
    else:
        doc["sale_price_idr"] = 0
    if "mrp_idr" in row:
        doc["mrp_idr"] = int(row.get("mrp_idr") or 0)
    elif existing:
        doc["mrp_idr"] = int(existing.get("mrp_idr") or 0)
    else:
        doc["mrp_idr"] = 0
    if existing:
        doc["created_at"] = existing.get("created_at")
        doc["created_by"] = existing.get("created_by")
        doc["id"] = existing["id"]
    else:
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = iso(now_utc())
        doc["created_by"] = user_id
    return doc
