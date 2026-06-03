"""Excel/CSV import/export for clinic treatments catalog."""
from __future__ import annotations

import csv
import io
import re
import uuid
from typing import Any, Dict, List, Optional, Tuple

from package_io import format_price_export, parse_price_idr
from saas import iso, now_utc

EXPORT_COLUMNS = [
    "ServiceCode",
    "ServiceName",
    "Category",
    "ServiceType",
    "ServicePrice",
    "OnlineBooking",
    "TaxIncluded",
    "TaxGroup",
    "ServiceLength",
]

HEADER_ALIASES = {
    "servicecode": "service_code",
    "service name": "name",
    "servicename": "name",
    "category": "category",
    "sub category": "sub_category",
    "subcategory": "sub_category",
    "businessunitname": "business_unit",
    "business unit": "business_unit",
    "servicetype": "service_type",
    "service type": "service_type",
    "serviceprice": "price",
    "service price": "price",
    "price": "price",
    "price idr": "price",
    "harga": "price",
    "onlinebooking": "online_booking",
    "online booking": "online_booking",
    "taxincluded": "tax_included",
    "tax included": "tax_included",
    "taxgroup": "tax_group",
    "tax group": "tax_group",
    "servicelength": "duration_min",
    "service length": "duration_min",
    "duration": "duration_min",
    "duration min": "duration_min",
}


def _norm_header(h: str) -> str:
    s = (h or "").replace("\ufeff", "").strip()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


def _resolve_header(norm: str) -> Optional[str]:
    if not norm:
        return None
    if norm in HEADER_ALIASES:
        return HEADER_ALIASES[norm]
    if "service" in norm and "price" in norm:
        return "price"
    if "service" in norm and "code" in norm:
        return "service_code"
    if "service" in norm and "name" in norm:
        return "name"
    if "service" in norm and "type" in norm:
        return "service_type"
    if "service" in norm and "length" in norm:
        return "duration_min"
    if "sub" in norm and "categ" in norm:
        return "sub_category"
    if "business" in norm and "unit" in norm:
        return "business_unit"
    if "online" in norm and "book" in norm:
        return "online_booking"
    if "tax" in norm and "incl" in norm:
        return "tax_included"
    if "tax" in norm and "group" in norm:
        return "tax_group"
    if norm == "price" or norm.endswith(" price"):
        return "price"
    return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def parse_bool(val: Any, default: bool = True) -> bool:
    if val is None or str(val).strip() == "":
        return default
    s = str(val).strip().lower()
    if s in ("true", "yes", "y", "1", "on"):
        return True
    if s in ("false", "no", "n", "0", "off"):
        return False
    return default


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def normalize_code(code: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "", (code or "")).lower()


def code_from_name(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "", name or "")
    return s[:32] if s else str(uuid.uuid4())[:8]


def build_treatment_lookup(existing_rows: List[dict]) -> Dict[str, dict]:
    """Index treatments by normalized code, name, and name-derived slug for import matching."""
    lookup: Dict[str, dict] = {}
    for t in existing_rows:
        for code in (t.get("service_code"), t.get("key")):
            nk = normalize_code(code or "")
            if nk:
                lookup[f"code:{nk}"] = t
        nm = normalize_name(t.get("name") or "")
        if nm:
            lookup[f"name:{nm}"] = t
        slug = normalize_code(t.get("name") or "")
        if slug:
            lookup[f"slug:{slug}"] = t
    return lookup


def find_treatment_match(row: dict, lookup: Dict[str, dict]) -> Optional[dict]:
    """Find existing treatment: service code first, then exact name, then name slug."""
    code = normalize_code(row.get("service_code") or "")
    if code:
        hit = lookup.get(f"code:{code}")
        if hit:
            return hit
    name = normalize_name(row.get("name") or "")
    if name:
        hit = lookup.get(f"name:{name}")
        if hit:
            return hit
    slug = normalize_code(row.get("name") or "")
    if slug:
        return lookup.get(f"slug:{slug}")
    return None


def register_treatment_in_lookup(lookup: Dict[str, dict], doc: dict) -> None:
    for code in (doc.get("service_code"), doc.get("key")):
        nk = normalize_code(code or "")
        if nk:
            lookup[f"code:{nk}"] = doc
    nm = normalize_name(doc.get("name") or "")
    if nm:
        lookup[f"name:{nm}"] = doc
    slug = normalize_code(doc.get("name") or "")
    if slug:
        lookup[f"slug:{slug}"] = doc


def parse_int(val: Any, default: int = 30) -> int:
    if val is None or str(val).strip() == "":
        return default
    try:
        return max(5, int(float(str(val).strip())))
    except (TypeError, ValueError):
        return default


def treatment_to_export_row(t: dict) -> Dict[str, Any]:
    return {
        "ServiceCode": t.get("service_code") or t.get("key") or "",
        "ServiceName": t.get("name") or "",
        "Category": t.get("category") or "",
        "ServiceType": t.get("service_type") or "None",
        "ServicePrice": format_price_export(t.get("price_idr") or 0),
        "OnlineBooking": "True" if t.get("active", True) else "False",
        "TaxIncluded": "True" if t.get("tax_included", True) else "False",
        "TaxGroup": t.get("tax_group") or "",
        "ServiceLength": int(t.get("duration_min") or 30),
    }


def rows_to_csv(rows: List[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def rows_to_xlsx(rows: List[dict]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel export") from ex
    wb = Workbook()
    ws = wb.active
    ws.title = "Treatments"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in EXPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def map_csv_headers(fieldnames: List[str]) -> Dict[str, str]:
    """Map CSV header -> internal field name."""
    out: Dict[str, str] = {}
    for raw in fieldnames or []:
        key = _resolve_header(_norm_header(raw))
        if key:
            out[raw] = key
    return out


def _map_columns_indexed(headers: List[Any]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for j, raw in enumerate(headers):
        internal = _resolve_header(_norm_header(_cell_str(raw)))
        if internal:
            out[j] = internal
    return out


def _find_header_row(rows: List[tuple]) -> Tuple[int, List[Any]]:
    for idx, row in enumerate(rows[:25]):
        mapped = _map_columns_indexed(list(row))
        if "name" in mapped.values():
            return idx, list(row)
    return 0, list(rows[0]) if rows else []


def _rows_from_table(all_rows: List[tuple]) -> Tuple[List[dict], List[dict]]:
    if not all_rows:
        return [], [{"row": 0, "message": "File has no rows"}]
    header_idx, headers = _find_header_row(all_rows)
    col_map = _map_columns_indexed(headers)
    if "name" not in col_map.values():
        preview = ", ".join(_cell_str(h) for h in headers if _cell_str(h))[:120]
        return [], [{"row": header_idx + 1, "message": f"Missing ServiceName column. Found: {preview or '(empty)'}"}]

    internal_fields = set(col_map.values())
    has_price_col = "price" in internal_fields
    parsed: List[dict] = []
    errors: List[dict] = []
    for i, raw_row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        if not raw_row or not any(_cell_str(c) for c in raw_row):
            continue
        data: Dict[str, Any] = {}
        row_list = list(raw_row)
        for j, internal in col_map.items():
            data[internal] = row_list[j] if j < len(row_list) else ""

        name = _cell_str(data.get("name"))
        if not name:
            errors.append({"row": i, "message": "ServiceName is required"})
            continue

        raw_code = _cell_str(data.get("service_code"))
        service_code = raw_code or code_from_name(name)

        row_out: Dict[str, Any] = {
            "service_code": service_code,
            "service_code_provided": bool(raw_code),
            "name": name,
            "category": _cell_str(data.get("category")) or "general",
            "service_type": _cell_str(data.get("service_type")) or "None",
            "active": parse_bool(data.get("online_booking"), True),
            "tax_included": parse_bool(data.get("tax_included"), True),
            "tax_group": _cell_str(data.get("tax_group")),
            "duration_min": parse_int(data.get("duration_min"), 30),
        }
        if "sub_category" in internal_fields:
            row_out["sub_category"] = _cell_str(data.get("sub_category"))
        if "business_unit" in internal_fields:
            row_out["business_unit"] = _cell_str(data.get("business_unit")) or "Default"
        if has_price_col:
            row_out["price_idr"] = parse_price_idr(data.get("price"))
        parsed.append(row_out)
    return parsed, errors


def parse_csv_text(text: str) -> Tuple[List[dict], List[dict]]:
    """Return (parsed_rows, errors)."""
    text = text.lstrip("\ufeff")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [{"row": 0, "message": "File has no header row"}]
    return _rows_from_table([tuple(r) for r in rows])


def parse_xlsx_bytes(raw: bytes) -> Tuple[List[dict], List[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as ex:
        raise RuntimeError("openpyxl is required for Excel import") from ex
    if raw[:2] != b"PK":
        return [], [{"row": 0, "message": "Not a valid .xlsx file. Save as Excel Workbook (.xlsx)"}]
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    last_errors: List[dict] = [{"row": 0, "message": "No treatment rows found"}]
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parsed, errors = _rows_from_table(rows)
        if parsed:
            wb.close()
            return parsed, errors
        if errors:
            last_errors = errors
    wb.close()
    return [], last_errors


def build_treatment_doc(row: dict, clinic_id: str, user_id: str, existing: Optional[dict] = None) -> dict:
    """Build MongoDB treatment document; performer_type preserved on update."""
    if existing and not row.get("service_code_provided"):
        service_code = (existing.get("service_code") or row.get("service_code") or "").strip()
    else:
        service_code = (row.get("service_code") or "").strip()
    if not service_code:
        service_code = code_from_name(row["name"])
    key = (existing.get("key") if existing else None) or service_code[:32]
    doc = {
        "name": row["name"],
        "key": key,
        "category": row.get("category") or "general",
        "service_type": row.get("service_type") or "None",
        "service_code": service_code,
        "active": bool(row.get("active", True)),
        "tax_included": bool(row.get("tax_included", True)),
        "tax_group": row.get("tax_group") or "",
        "duration_min": int(row.get("duration_min") or 30),
        "clinic_id": clinic_id,
    }
    if "sub_category" in row:
        doc["sub_category"] = row.get("sub_category") or ""
    elif existing:
        doc["sub_category"] = existing.get("sub_category") or ""
    else:
        doc["sub_category"] = ""

    if "business_unit" in row:
        doc["business_unit"] = row.get("business_unit") or "Default"
    elif existing:
        doc["business_unit"] = existing.get("business_unit") or "Default"
    else:
        doc["business_unit"] = "Default"

    if existing:
        doc["performer_type"] = existing.get("performer_type") or "therapist"
        doc["allowed_performer_roles"] = existing.get("allowed_performer_roles")
        doc["allow_multiple_performers"] = existing.get("allow_multiple_performers", False)
        doc["requires_assistant"] = existing.get("requires_assistant", False)
        doc["consent_required"] = existing.get("consent_required", False)
        if "price_idr" in row:
            doc["price_idr"] = int(row.get("price_idr") or 0)
        else:
            doc["price_idr"] = existing.get("price_idr", 0)
        doc["slots_per_session"] = existing.get("slots_per_session", 1)
        doc["description"] = existing.get("description") or ""
        doc["created_at"] = existing.get("created_at")
        doc["created_by"] = existing.get("created_by")
        doc["id"] = existing["id"]
    else:
        doc["performer_type"] = "therapist"
        doc["price_idr"] = int(row.get("price_idr") or 0)
        doc["slots_per_session"] = 1
        doc["description"] = ""
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = iso(now_utc())
        doc["created_by"] = user_id
    return doc
