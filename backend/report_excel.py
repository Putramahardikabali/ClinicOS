"""Excel (.xlsx) export helpers for clinic reports."""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

Row = Sequence[Any]
SheetSpec = Dict[str, Any]

HEADER_FILL = PatternFill("solid", fgColor="52796F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
META_FONT = Font(color="444444")


def _safe_sheet_name(name: str) -> str:
    s = re.sub(r'[\\/*?:\[\]]', " ", (name or "Sheet"))[:31].strip()
    return s or "Sheet"


def _auto_width(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(48, max(10, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[letter].width = width


def _write_table(ws, start_row: int, headers: List[str], rows: List[Row]) -> int:
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r = start_row + 1
    for row in rows:
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
        r += 1
    _auto_width(ws)
    return r


def build_report_workbook(
    title: str,
    range_info: dict,
    filters: Optional[Dict[str, Any]] = None,
    summary_rows: Optional[List[Tuple[str, Any]]] = None,
    detail_sheets: Optional[List[SheetSpec]] = None,
) -> bytes:
    """Build xlsx with Info + Summary + optional detail sheets."""
    wb = Workbook()
    info = wb.active
    info.title = "Info"

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    preset = (range_info or {}).get("preset") or "custom"
    from_d = (range_info or {}).get("from") or "—"
    to_d = (range_info or {}).get("to") or "—"

    info["A1"] = title
    info["A1"].font = TITLE_FONT
    meta = [
        ("Report", title),
        ("Date range", f"{from_d} to {to_d}"),
        ("Preset", preset),
        ("Generated", generated),
    ]
    flt = filters or {}
    for k, v in sorted(flt.items()):
        if v is not None and v != "" and v != "all":
            meta.append((f"Filter: {k}", v))

    row = 3
    for label, val in meta:
        info.cell(row=row, column=1, value=label).font = Font(bold=True)
        info.cell(row=row, column=2, value=val)
        row += 1
    _auto_width(info)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Summary"
    summary["A1"].font = TITLE_FONT
    rows = summary_rows or []
    if rows:
        _write_table(summary, 3, ["Metric", "Value"], [[k, v] for k, v in rows])
    else:
        summary["A3"] = "No summary data."

    for spec in detail_sheets or []:
        name = _safe_sheet_name(spec.get("name") or "Detail")
        ws = wb.create_sheet(name)
        ws["A1"] = spec.get("title") or name
        ws["A1"].font = TITLE_FONT
        headers = spec.get("headers") or []
        data_rows = spec.get("rows") or []
        if headers:
            _write_table(ws, 3, headers, data_rows)
        else:
            ws["A3"] = "No detail rows."

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_response(filename: str, content: bytes) -> Response:
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def summary_from_dict(d: dict, labels: Optional[Dict[str, str]] = None) -> List[Tuple[str, Any]]:
    out: List[Tuple[str, Any]] = []
    labels = labels or {}
    for k, v in d.items():
        out.append((labels.get(k, k.replace("_", " ").title()), v))
    return out


def dict_rows(items: List[dict], columns: List[Tuple[str, str]]) -> Tuple[List[str], List[Row]]:
    headers = [c[1] for c in columns]
    rows: List[Row] = []
    for item in items:
        rows.append([item.get(c[0]) for c in columns])
    return headers, rows


def kv_rows(items: List[dict], key: str, val: str, key_label: str, val_label: str) -> Tuple[List[str], List[Row]]:
    return [key_label, val_label], [[i.get(key), i.get(val)] for i in items]
