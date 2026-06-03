"""Super Admin Phase 3: platform reliability, monitoring, support, exports."""
from __future__ import annotations

import io
import uuid
import zipfile
import csv
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from saas import iso, now_utc, resolve_clinic_limits, PLAN_CATALOG as DEFAULT_PLAN_CATALOG
from platform_ops import create_platform_notification

_SENSITIVE = frozenset({
    "password", "password_hash", "token", "authorization", "secret", "signature",
    "clinical", "diagnosis", "note", "body", "content",
})


def _sanitize_text(text: str, max_len: int = 500) -> str:
    if not text:
        return ""
    s = str(text)
    for key in _SENSITIVE:
        if key in s.lower():
            return "[redacted — sensitive content omitted]"
    if len(s) > max_len:
        return s[:max_len] + "…"
    return s


async def log_platform_error(
    db,
    *,
    module: str,
    message: str,
    severity: str = "error",
    clinic_id: Optional[str] = None,
    clinic_name: Optional[str] = None,
    error_type: str = "platform",
    meta: Optional[dict] = None,
) -> str:
    doc = {
        "id": str(uuid.uuid4()),
        "type": error_type,
        "module": module or "unknown",
        "message": _sanitize_text(message),
        "severity": severity if severity in ("info", "warning", "error", "critical") else "error",
        "status": "open",
        "clinic_id": clinic_id,
        "clinic_name": clinic_name,
        "meta": {k: _sanitize_text(str(v), 200) for k, v in (meta or {}).items() if k not in _SENSITIVE},
        "created_at": iso(now_utc()),
        "resolved_at": None,
        "resolved_by": None,
    }
    await db.platform_errors.insert_one(doc)
    return doc["id"]


async def log_failed_login(db, email: str, reason: str = "invalid_credentials", clinic_id: Optional[str] = None):
    clinic_name = None
    if clinic_id:
        c = await db.clinics.find_one({"id": clinic_id}, {"_id": 0, "name": 1})
        clinic_name = (c or {}).get("name")
    await log_platform_error(
        db,
        module="auth",
        message=f"Failed login attempt for {email[:3]}***",
        severity="warning",
        clinic_id=clinic_id,
        clinic_name=clinic_name,
        error_type="failed_login",
        meta={"reason": reason},
    )


async def _mongo_ok(db) -> dict:
    try:
        await db.command("ping")
        return {"status": "ok", "message": "Connected"}
    except Exception as e:
        return {"status": "error", "message": _sanitize_text(str(e), 120)}


async def _storage_ok(init_storage) -> dict:
    if not init_storage:
        return {"status": "unknown", "message": "Storage not configured"}
    try:
        init_storage()
        return {"status": "ok", "message": "Storage API reachable"}
    except Exception as e:
        return {"status": "error", "message": _sanitize_text(str(e), 120)}


async def build_system_health(db, *, init_storage=None) -> dict:
    now = now_utc()
    since_24h = iso(now - timedelta(hours=24))
    open_errors = await db.platform_errors.count_documents({"status": "open"})
    recent_errors = await db.platform_errors.find(
        {"created_at": {"$gte": since_24h}},
        {"_id": 0},
    ).sort("created_at", -1).limit(10).to_list(10)
    failed_uploads = await db.platform_errors.count_documents({
        "type": "failed_upload",
        "created_at": {"$gte": since_24h},
    })
    failed_sends = await db.platform_errors.count_documents({
        "type": "failed_send",
        "created_at": {"$gte": since_24h},
    })
    failed_logins = await db.platform_errors.count_documents({
        "type": "failed_login",
        "created_at": {"$gte": since_24h},
    })
    pending_payments = await db.payment_requests.count_documents({"status": "submitted"})
    backup_status = await db.platform_backup_status.find_one({"id": "current"}, {"_id": 0}) or {}
    failed_backups = await db.platform_backups.count_documents({"status": "failed", "started_at": {"$gte": since_24h}})
    return {
        "checked_at": iso(now),
        "backend": {"status": "ok", "message": "API process running"},
        "database": await _mongo_ok(db),
        "storage": await _storage_ok(init_storage),
        "queue_jobs": {
            "status": "ok",
            "pending_payments": pending_payments,
            "open_support_requests": await db.platform_support_requests.count_documents({"status": {"$in": ["open", "in_progress"]}}),
            "message": "No background worker — payment queue shown as proxy",
        },
        "backups": {
            "last_database": backup_status.get("last_db_backup_at"),
            "last_files": backup_status.get("last_file_backup_at"),
            "db_status": backup_status.get("last_db_status", "unknown"),
            "file_status": backup_status.get("last_file_status", "unknown"),
            "failed_last_24h": failed_backups,
        },
        "failed_uploads_24h": failed_uploads,
        "failed_sends_24h": failed_sends,
        "failed_logins_24h": failed_logins,
        "open_errors": open_errors,
        "recent_errors": recent_errors,
    }


def _activity_label(last_any: Optional[datetime], sub_status: str) -> str:
    if sub_status in ("archived", "cancelled", "suspended"):
        return "Inactive"
    now = now_utc()
    if not last_any:
        return "At risk"
    days = (now - last_any).days
    if days <= 7:
        return "Active"
    if days <= 30:
        return "Quiet"
    return "At risk" if days <= 90 else "Inactive"


async def clinic_activity_row(db, clinic: dict) -> dict:
    cid = clinic["id"]
    since_7d = iso(now_utc() - timedelta(days=7))

    async def _latest(coll, date_field="created_at"):
        row = await db[coll].find_one({"clinic_id": cid}, {"_id": 0, date_field: 1}, sort=[(date_field, -1)])
        return (row or {}).get(date_field)

    last_login_row = await db.audit_logs.find_one(
        {"clinic_id": cid, "action": "login"},
        {"_id": 0, "created_at": 1},
        sort=[("created_at", -1)],
    )
    owner = await db.users.find_one({"clinic_id": cid, "role": "super_admin"}, {"_id": 0, "last_login_at": 1})
    last_login = (owner or {}).get("last_login_at") or (last_login_row or {}).get("created_at")
    last_booking = await _latest("bookings")
    last_visit = await _latest("visits")
    last_invoice = await _latest("invoices")
    active_users_7d = await db.audit_logs.distinct(
        "user_id",
        {"clinic_id": cid, "action": "login", "created_at": {"$gte": since_7d}},
    )

    dates = []
    for d in (last_login, last_booking, last_visit, last_invoice):
        if d:
            try:
                dates.append(datetime.fromisoformat(d.replace("Z", "+00:00") if isinstance(d, str) else d))
            except Exception:
                pass
    last_any = max(dates) if dates else None
    sub = clinic.get("subscription") or {}

    return {
        "clinic_id": cid,
        "clinic_name": clinic.get("name"),
        "slug": clinic.get("slug"),
        "plan": sub.get("plan"),
        "status": sub.get("status"),
        "last_login": last_login,
        "last_booking": last_booking,
        "last_visit": last_visit,
        "last_invoice": last_invoice,
        "active_users_7d": len(active_users_7d),
        "activity_label": _activity_label(last_any, sub.get("status", "trial")),
    }


async def usage_analytics_for_clinic(db, clinic: dict, plan_catalog: dict) -> dict:
    cid = clinic["id"]
    usage = {
        "staff_count": await db.users.count_documents({"clinic_id": cid}),
        "patient_count": await db.patients.count_documents({"clinic_id": cid}),
        "visit_count": await db.visits.count_documents({"clinic_id": cid}),
        "invoice_count": await db.invoices.count_documents({"clinic_id": cid}),
        "file_count": await db.photos.count_documents({"clinic_id": cid}),
    }
    agg = await db.photos.aggregate([
        {"$match": {"clinic_id": cid}},
        {"$group": {"_id": None, "bytes": {"$sum": {"$ifNull": ["$size_bytes", 0]}}}},
    ]).to_list(1)
    storage_bytes = int((agg[0]["bytes"] if agg else 0) or 0)
    usage["storage_used_gb"] = round(storage_bytes / (1024 ** 3), 2)
    limits = resolve_clinic_limits(clinic)
    max_staff = int(limits.get("max_staff") or 3)
    max_storage = int(limits.get("storage_gb") or 2)
    alerts = []
    for pct, label in ((80, "warning"), (95, "critical")):
        if max_staff < 9999 and usage["staff_count"] >= max_staff * pct / 100:
            alerts.append({"metric": "staff", "level": label, "used": usage["staff_count"], "limit": max_staff})
        if usage["storage_used_gb"] >= max_storage * pct / 100:
            alerts.append({"metric": "storage", "level": label, "used": usage["storage_used_gb"], "limit": max_storage})
    return {
        "clinic_id": cid,
        "clinic_name": clinic.get("name"),
        "slug": clinic.get("slug"),
        "usage": usage,
        "limits": {"max_staff": max_staff, "storage_gb": max_storage},
        "alerts": alerts,
    }


async def scan_usage_limit_alerts(db, plan_catalog: dict):
    clinics = await db.clinics.find(
        {"subscription.status": {"$nin": ["archived", "cancelled"]}},
        {"_id": 0},
    ).to_list(2000)
    for c in clinics:
        row = await usage_analytics_for_clinic(db, c, plan_catalog)
        for al in row.get("alerts") or []:
            pct = 95 if al["level"] == "critical" else 80
            ntype = "staff_limit" if al["metric"] == "staff" else "storage_limit"
            await create_platform_notification(
                db,
                ntype=ntype,
                title=f"Usage alert ({al['level']}): {c.get('name')}",
                body=f"{al['metric']} at {al['used']} / {al['limit']} (≥{pct}%)",
                clinic_id=c["id"],
                clinic_name=c.get("name"),
                link=f"/superadmin/ops?tab=analytics",
            )


def _csv_bytes(headers: List[str], rows: List[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow({h: r.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8-sig")


_EXPORT_SKIP_FIELDS = frozenset({
    "password_hash", "invite_token", "token", "clinical_notes", "note_body", "notes",
})


def _export_row(row: dict) -> dict:
    return {k: v for k, v in row.items() if k not in _EXPORT_SKIP_FIELDS and not k.startswith("_")}


async def _clinic_export_tables(db, clinic_id: str) -> dict:
    c = await db.clinics.find_one({"id": clinic_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Clinic not found")
    clinic_meta = [{
        "id": c.get("id"), "name": c.get("name"), "slug": c.get("slug"),
        "email": c.get("email"), "owner_email": c.get("owner_email"),
        "plan": (c.get("subscription") or {}).get("plan"),
        "status": (c.get("subscription") or {}).get("status"),
        "created_at": c.get("created_at"),
    }]
    users = [_export_row(u) for u in await db.users.find({"clinic_id": clinic_id}, {"_id": 0, "password_hash": 0, "invite_token": 0}).to_list(500)]
    patients = [_export_row(p) for p in await db.patients.find({"clinic_id": clinic_id}, {"_id": 0}).to_list(10000)]
    tables = {"clinic": clinic_meta, "users": users, "patients": patients}
    for coll, fields in (
        ("bookings", ["id", "patient_id", "scheduled_at", "status", "created_at"]),
        ("visits", ["id", "patient_id", "status", "created_at"]),
        ("invoices", ["id", "patient_id", "payment_status", "total_idr", "created_at"]),
        ("treatments", ["id", "name", "price_idr", "active"]),
        ("packages", ["id", "name", "price_idr", "active"]),
        ("products", ["id", "name", "sku", "active"]),
        ("payment_requests", ["id", "plan", "amount_idr", "status", "created_at"]),
    ):
        rows = await db[coll].find({"clinic_id": clinic_id}, {"_id": 0}).to_list(5000)
        if rows:
            tables[coll] = [{k: r.get(k, "") for k in fields} for r in rows]
    return {"clinic": c, "tables": tables}


async def build_clinic_export_zip(db, clinic_id: str) -> bytes:
    bundle = await _clinic_export_tables(db, clinic_id)
    slug = bundle["clinic"].get("slug", clinic_id)
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, rows in bundle["tables"].items():
            if not rows:
                continue
            headers = list(rows[0].keys())
            zf.writestr(f"{name}.csv", _csv_bytes(headers, rows))
    mem.seek(0)
    return mem.read()


async def build_clinic_export_xlsx(db, clinic_id: str) -> bytes:
    from openpyxl import Workbook

    bundle = await _clinic_export_tables(db, clinic_id)
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in bundle["tables"].items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    if not wb.sheetnames:
        wb.create_sheet("clinic")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------- API models ----------------
class ErrorStatusIn(BaseModel):
    status: str  # resolved | ignored


class SupportRequestIn(BaseModel):
    clinic_id: str
    subject: str
    priority: str = "normal"
    assigned_to: Optional[str] = None
    internal_note: Optional[str] = None


class SupportUpdateIn(BaseModel):
    status: Optional[str] = None
    priority: Optional[str] = None
    assigned_to: Optional[str] = None
    internal_note: Optional[str] = None


class BackupRecordIn(BaseModel):
    backup_type: str  # database | files
    status: str  # success | failed | running
    message: Optional[str] = None
    size_bytes: Optional[int] = None


class ExportConfirmIn(BaseModel):
    confirm_slug: str
    reason: str


def register_platform_reliability(
    api: APIRouter,
    db,
    get_current_user,
    audit,
    *,
    init_storage=None,
    PLAN_CATALOG=None,
    invalidate_clinic_sessions=None,
    invalidate_user_sessions=None,
):
    plan_catalog = PLAN_CATALOG or DEFAULT_PLAN_CATALOG

    async def admin_dep(user: dict = Depends(get_current_user)):
        if not user.get("platform_admin"):
            raise HTTPException(status_code=403, detail="Platform admin only")
        return user

    async def _sa_audit(user, action, record_id="", meta=None, reason=None):
        await audit(
            {"id": "platform-admin", "email": user["email"], "role": "platform_admin", "clinic_id": None, "name": "Platform Admin"},
            action, "platform_ops", record_id, meta=meta, reason=reason,
        )

    @api.get("/superadmin/ops/health")
    async def ops_health(_user: dict = Depends(admin_dep)):
        await scan_usage_limit_alerts(db, plan_catalog)
        return await build_system_health(db, init_storage=init_storage)

    @api.get("/superadmin/ops/errors")
    async def ops_list_errors(status: Optional[str] = None, limit: int = 100, _user: dict = Depends(admin_dep)):
        flt: Dict[str, Any] = {}
        if status:
            flt["status"] = status
        cap = min(max(1, limit), 500)
        rows = await db.platform_errors.find(flt, {"_id": 0}).sort("created_at", -1).limit(cap).to_list(cap)
        return rows

    @api.put("/superadmin/ops/errors/{eid}/status")
    async def ops_update_error(eid: str, payload: ErrorStatusIn, user: dict = Depends(admin_dep)):
        if payload.status not in ("resolved", "ignored", "open"):
            raise HTTPException(status_code=400, detail="Invalid status")
        now = iso(now_utc())
        upd = {"status": payload.status}
        if payload.status in ("resolved", "ignored"):
            upd["resolved_at"] = now
            upd["resolved_by"] = user["email"]
        r = await db.platform_errors.update_one({"id": eid}, {"$set": upd})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Error not found")
        await _sa_audit(user, "error_status_updated", eid, meta={"status": payload.status})
        return await db.platform_errors.find_one({"id": eid}, {"_id": 0})

    @api.get("/superadmin/ops/backups")
    async def ops_backups(limit: int = 50, _user: dict = Depends(admin_dep)):
        status = await db.platform_backup_status.find_one({"id": "current"}, {"_id": 0}) or {}
        history = await db.platform_backups.find({}, {"_id": 0}).sort("started_at", -1).limit(min(limit, 200)).to_list(min(limit, 200))
        failed = await db.platform_backups.find({"status": "failed"}, {"_id": 0}).sort("started_at", -1).limit(20).to_list(20)
        return {"status": status, "history": history, "recent_failures": failed}

    @api.post("/superadmin/ops/backups/record")
    async def ops_record_backup(payload: BackupRecordIn, user: dict = Depends(admin_dep)):
        if payload.backup_type not in ("database", "files"):
            raise HTTPException(status_code=400, detail="Invalid backup type")
        if payload.status not in ("success", "failed", "running"):
            raise HTTPException(status_code=400, detail="Invalid status")
        now = iso(now_utc())
        doc = {
            "id": str(uuid.uuid4()),
            "backup_type": payload.backup_type,
            "status": payload.status,
            "message": _sanitize_text(payload.message or "", 300),
            "size_bytes": payload.size_bytes,
            "started_at": now,
            "completed_at": now if payload.status != "running" else None,
            "recorded_by": user["email"],
        }
        await db.platform_backups.insert_one(doc)
        status_upd: Dict[str, Any] = {"id": "current", "updated_at": now}
        if payload.backup_type == "database":
            status_upd["last_db_backup_at"] = now
            status_upd["last_db_status"] = payload.status
        else:
            status_upd["last_file_backup_at"] = now
            status_upd["last_file_status"] = payload.status
        await db.platform_backup_status.update_one({"id": "current"}, {"$set": status_upd}, upsert=True)
        if payload.status == "failed":
            await log_platform_error(db, module="backup", message=f"{payload.backup_type} backup failed", severity="critical", error_type="backup_failed", meta={"message": payload.message})
        await _sa_audit(user, "backup_recorded", doc["id"], meta={"type": payload.backup_type, "status": payload.status})
        doc.pop("_id", None)
        return doc

    @api.get("/superadmin/ops/support")
    async def ops_list_support(status: Optional[str] = None, clinic_id: Optional[str] = None, _user: dict = Depends(admin_dep)):
        flt: Dict[str, Any] = {}
        if status:
            flt["status"] = status
        if clinic_id:
            flt["clinic_id"] = clinic_id
        rows = await db.platform_support_requests.find(flt, {"_id": 0}).sort("updated_at", -1).limit(200).to_list(200)
        return rows

    @api.post("/superadmin/ops/support")
    async def ops_create_support(payload: SupportRequestIn, user: dict = Depends(admin_dep)):
        c = await db.clinics.find_one({"id": payload.clinic_id}, {"_id": 0, "name": 1})
        if not c:
            raise HTTPException(status_code=404, detail="Clinic not found")
        now = iso(now_utc())
        notes = []
        if payload.internal_note:
            notes.append({"at": now, "by": user["email"], "text": payload.internal_note.strip()})
        doc = {
            "id": str(uuid.uuid4()),
            "clinic_id": payload.clinic_id,
            "clinic_name": c.get("name"),
            "subject": payload.subject.strip(),
            "priority": payload.priority if payload.priority in ("low", "normal", "high", "urgent") else "normal",
            "status": "open",
            "assigned_to": payload.assigned_to,
            "internal_notes": notes,
            "created_at": now,
            "updated_at": now,
            "created_by": user["email"],
        }
        await db.platform_support_requests.insert_one(doc)
        await _sa_audit(user, "support_request_created", doc["id"], meta={"clinic_id": payload.clinic_id, "subject": doc["subject"]})
        doc.pop("_id", None)
        return doc

    @api.put("/superadmin/ops/support/{sid}")
    async def ops_update_support(sid: str, payload: SupportUpdateIn, user: dict = Depends(admin_dep)):
        req = await db.platform_support_requests.find_one({"id": sid}, {"_id": 0})
        if not req:
            raise HTTPException(status_code=404, detail="Support request not found")
        upd: Dict[str, Any] = {"updated_at": iso(now_utc())}
        if payload.status:
            if payload.status not in ("open", "in_progress", "resolved", "closed"):
                raise HTTPException(status_code=400, detail="Invalid status")
            upd["status"] = payload.status
        if payload.priority:
            upd["priority"] = payload.priority
        if payload.assigned_to is not None:
            upd["assigned_to"] = payload.assigned_to
        if payload.internal_note:
            notes = list(req.get("internal_notes") or [])
            notes.append({"at": iso(now_utc()), "by": user["email"], "text": payload.internal_note.strip()})
            upd["internal_notes"] = notes
        await db.platform_support_requests.update_one({"id": sid}, {"$set": upd})
        await _sa_audit(user, "support_request_updated", sid, meta=upd)
        return await db.platform_support_requests.find_one({"id": sid}, {"_id": 0})

    @api.get("/superadmin/ops/activity")
    async def ops_activity(list_filter: Optional[str] = None, _user: dict = Depends(admin_dep)):
        flt: Dict[str, Any] = {"subscription.status": {"$ne": "archived"}}
        if list_filter in ("Active", "Quiet", "At risk", "Inactive"):
            pass  # filter post-compute
        clinics = await db.clinics.find(flt, {"_id": 0}).sort("name", 1).to_list(2000)
        rows = []
        for c in clinics:
            row = await clinic_activity_row(db, c)
            if list_filter and row["activity_label"] != list_filter:
                continue
            rows.append(row)
        summary = {}
        for r in rows:
            summary[r["activity_label"]] = summary.get(r["activity_label"], 0) + 1
        return {"summary": summary, "clinics": rows}

    @api.get("/superadmin/ops/analytics")
    async def ops_analytics(_user: dict = Depends(admin_dep)):
        clinics = await db.clinics.find({"subscription.status": {"$ne": "archived"}}, {"_id": 0}).to_list(2000)
        rows = [await usage_analytics_for_clinic(db, c, plan_catalog) for c in clinics]
        alert_count = sum(1 for r in rows if r.get("alerts"))
        return {"alert_clinics": alert_count, "clinics": rows}

    @api.get("/superadmin/ops/security")
    async def ops_security(limit: int = 100, _user: dict = Depends(admin_dep)):
        since = iso(now_utc() - timedelta(days=7))
        failed_logins = await db.platform_errors.find(
            {"type": "failed_login", "created_at": {"$gte": since}},
            {"_id": 0},
        ).sort("created_at", -1).limit(min(limit, 200)).to_list(min(limit, 200))
        recent_logins = await db.audit_logs.find(
            {"action": "login"},
            {"_id": 0, "user_email": 1, "clinic_id": 1, "created_at": 1},
        ).sort("created_at", -1).limit(50).to_list(50)
        for r in recent_logins:
            if r.get("clinic_id"):
                c = await db.clinics.find_one({"id": r["clinic_id"]}, {"_id": 0, "name": 1})
                r["clinic_name"] = (c or {}).get("name")
        return {
            "failed_logins_7d": len(failed_logins),
            "failed_logins": failed_logins,
            "recent_logins": recent_logins,
            "active_sessions_note": "Session invalidation uses auth_version; no per-session listing stored.",
        }

    @api.post("/superadmin/clinics/{cid}/export")
    async def ops_export_clinic(
        cid: str,
        payload: ExportConfirmIn,
        format: str = Query("zip", description="zip or xlsx"),
        user: dict = Depends(admin_dep),
    ):
        reason = (payload.reason or "").strip()
        if not reason:
            raise HTTPException(status_code=400, detail="Export reason is required")
        fmt = (format or "zip").lower()
        if fmt not in ("zip", "xlsx"):
            raise HTTPException(status_code=400, detail="Format must be zip or xlsx")
        c = await db.clinics.find_one({"id": cid}, {"_id": 0, "name": 1, "slug": 1})
        if not c:
            raise HTTPException(status_code=404, detail="Clinic not found")
        if payload.confirm_slug.strip() != c.get("slug"):
            raise HTTPException(status_code=400, detail="Clinic slug confirmation does not match")
        await audit(
            {"id": "platform-admin", "email": user["email"], "role": "platform_admin", "clinic_id": cid, "name": "Platform Admin"},
            "clinic_support_data_exported",
            "clinic",
            cid,
            meta={
                "clinic_name": c.get("name"),
                "slug": c.get("slug"),
                "format": fmt,
                "export_type": "support_data",
                "exported_by": user["email"],
            },
            reason=reason,
        )
        slug = c.get("slug", cid)
        if fmt == "xlsx":
            data = await build_clinic_export_xlsx(db, cid)
            filename = f"clinicos-support-data-{slug}.xlsx"
            media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            data = await build_clinic_export_zip(db, cid)
            filename = f"clinicos-support-data-{slug}.zip"
            media = "application/zip"
        return StreamingResponse(
            io.BytesIO(data),
            media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
