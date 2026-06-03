"""Patient import/export helpers and API tests."""
import io
import uuid

import pytest

from patient_io import (
    EXPORT_COLUMNS,
    format_last_visit_export,
    parse_csv_text,
    parse_last_visit,
    patient_to_export_row,
    rows_to_xlsx,
)


def test_parse_last_visit_dd_mm_yyyy():
    assert parse_last_visit("21/01/2026") == "2026-01-21"
    assert parse_last_visit("14/10/2022") == "2022-10-14"
    assert parse_last_visit("") == ""


def test_export_row_roundtrip_fields():
    row = patient_to_export_row({
        "first_name": "Setareh",
        "last_name": "Torney",
        "full_name": "Setareh Torney",
        "phone": "1121215258",
        "user_code": "BL21231",
        "membership_name": "",
        "last_visit": "2026-01-21",
        "guest_icon_information": "",
    })
    assert row["FirstName"] == "Setareh"
    assert row["LastName"] == "Torney"
    assert row["UserCode"] == "BL21231"
    assert row["lastvisit"] == "21/01/2026"


def test_parse_csv_minimal():
    csv_body = (
        "FirstName,LastName,Phone No,UserCode,membershipname,lastvisit,guestIconInformation\n"
        "Test,Import,081299988877,BL99999,,21/01/2026,\n"
    )
    parsed, errors = parse_csv_text(csv_body)
    assert not errors
    assert len(parsed) == 1
    assert parsed[0]["full_name"] == "Test Import"
    assert parsed[0]["user_code"] == "BL99999"
    assert parsed[0]["last_visit"] == "2026-01-21"


def test_parse_csv_with_title_row():
    csv_body = (
        "Patient export\n"
        "FirstName,LastName,Phone No,UserCode,membershipname,lastvisit,guestIconInformation\n"
        "Jane,Doe,08123456789,BL12345,,,\n"
    )
    parsed, errors = parse_csv_text(csv_body)
    assert not errors or len(parsed) == 1
    assert len(parsed) == 1
    assert parsed[0]["full_name"] == "Jane Doe"


def test_parse_headers_with_extra_spaces():
    csv_body = " FirstName , LastName , Phone No \nA,B,0812\n"
    parsed, errors = parse_csv_text(csv_body)
    assert len(parsed) == 1


def test_rows_to_xlsx_has_headers():
    data = rows_to_xlsx([])
    assert len(data) > 100
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    assert headers == EXPORT_COLUMNS
