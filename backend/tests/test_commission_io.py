"""Commission export helpers."""
from commission_io import (
    DETAIL_COLUMNS,
    STAFF_SUMMARY_COLUMNS,
    build_staff_period_summary,
    commission_record_to_detail_row,
    commission_record_to_export_row,
    month_to_range,
    rows_to_staff_export_xlsx,
    rows_to_xlsx,
)


def test_date_range_validation_logic():
    assert month_to_range("2026-05") == ("2026-05-01", "2026-05-31")
    assert "2026-05-01" <= "2026-05-15" <= "2026-05-31"


def test_commission_export_row():
    row = commission_record_to_export_row({
        "approved_at": "2026-05-15T10:00:00+00:00",
        "invoice_number_snapshot": "INV-20260515-0001",
        "item_name_snapshot": "Facial",
        "item_type": "treatment",
        "staff_name_snapshot": "Dr A",
        "staff_role_snapshot": "doctor",
        "gross_amount": 500000,
        "discount_amount": 0,
        "net_amount": 500000,
        "paid_amount": 500000,
        "commission_rule_name_snapshot": "10% treatments",
        "commission_type": "percentage",
        "commission_value": 10,
        "calculation_basis": "paid",
        "commission_amount": 50000,
        "status": "approved",
    }, "Jane Doe")
    assert row["PatientName"] == "Jane Doe"
    assert row["CommissionAmount"] == 50000
    assert row["CommissionValue"] == "10%"


def test_staff_detail_row():
    row = commission_record_to_detail_row({
        "approved_at": "2026-05-15T10:00:00+00:00",
        "created_at": "2026-05-14T10:00:00+00:00",
        "invoice_number_snapshot": "INV-1",
        "item_name_snapshot": "Facial",
        "gross_amount": 100,
        "discount_amount": 10,
        "net_amount": 90,
        "paid_amount": 90,
        "commission_rule_name_snapshot": "Rule A",
        "commission_amount": 9,
        "status": "approved",
    }, "Jane", date_basis="approved_at")
    assert row["Patient"] == "Jane"
    assert row["Commission Amount"] == 9


def test_staff_period_summary():
    summary = build_staff_period_summary(
        [
            {"gross_amount": 100, "discount_amount": 0, "net_amount": 100, "paid_amount": 100, "commission_amount": 10, "status": "approved"},
            {"gross_amount": 200, "discount_amount": 20, "net_amount": 180, "paid_amount": 180, "commission_amount": 18, "status": "paid_out"},
        ],
        staff_name="Dr A",
        staff_role="doctor",
        period_start="2026-05-01",
        period_end="2026-05-31",
    )
    assert summary["Staff Name"] == "Dr A"
    assert summary["Total Commission"] == 28
    assert summary["Approved Count"] == 1
    assert summary["Paid Out Count"] == 1
    assert summary["Remaining Approved Unpaid"] == 10


def test_rows_to_staff_export_xlsx_has_sheets():
    data = rows_to_staff_export_xlsx(
        build_staff_period_summary([], staff_name="Dr A", staff_role="doctor", period_start="2026-05-01", period_end="2026-05-31"),
        [],
    )
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames[0] == "Summary"
    assert "Detailed Commission Items" in wb.sheetnames
    summary_headers = [c.value for c in wb["Summary"][1]]
    assert summary_headers == STAFF_SUMMARY_COLUMNS
    detail_headers = [c.value for c in wb["Detailed Commission Items"][1]]
    assert detail_headers == DETAIL_COLUMNS


def test_rows_to_xlsx_legacy_has_headers():
    from commission_io import EXPORT_COLUMNS
    data = rows_to_xlsx([])
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == EXPORT_COLUMNS
